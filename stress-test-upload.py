#!/usr/bin/env python3
"""
SMG Bridge — Upload Stress Tester
===================================
Generates realistic patient Excel files and uploads them to the portal.

Usage:
  python3 stress-test-upload.py                           # 20 files, 50 rows each (requires BRIDGE_AUTH_ENABLED=false on server)
  python3 stress-test-upload.py --user admin --pass smgadmin  # authenticate first
  python3 stress-test-upload.py --files 100               # 100 files
  python3 stress-test-upload.py --files 50 --rows 200     # 50 files, 200 rows each
  python3 stress-test-upload.py --files 10 --batch 5      # upload in batches of 5
  python3 stress-test-upload.py --url http://localhost:3001   # custom server URL
  python3 stress-test-upload.py --types patients,claims,pharmacy  # specific sheet types
  python3 stress-test-upload.py --keep                    # keep generated files after upload
"""

import argparse
import csv
import io
import json
import os
import random
import sys
import time
from datetime import date, timedelta
from pathlib import Path

# ── Try openpyxl (best), fall back to xlsxwriter, fall back to CSV ─────────────
try:
    import openpyxl
    EXCEL_ENGINE = "openpyxl"
except ImportError:
    try:
        import xlsxwriter
        EXCEL_ENGINE = "xlsxwriter"
    except ImportError:
        EXCEL_ENGINE = "csv"
        print("[WARN] openpyxl/xlsxwriter not found — generating CSV files instead.")
        print("       Install openpyxl for Excel:  pip install openpyxl\n")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("[WARN] requests not installed. Run:  pip install requests")
    print("       Files will be generated but NOT uploaded.\n")

random.seed()  # non-reproducible for stress variety

# ── Realistic data pools ──────────────────────────────────────────────────────

KOREAN_SURNAMES = ["KIM","LEE","PARK","CHOI","JUNG","KANG","CHO","CHUNG","SHIN",
                   "YOO","YOON","HAN","OH","SEO","HWANG","AHN","SONG","MOON","YANG","BAE"]
NON_KOREAN_SURNAMES = ["GARCIA","MARTINEZ","HERNANDEZ","LOPEZ","GONZALEZ","NGUYEN",
                        "TRAN","SMITH","JOHNSON","WILLIAMS","BROWN","JONES","CRUZ","REYES"]
ALL_SURNAMES = KOREAN_SURNAMES + NON_KOREAN_SURNAMES

FIRST_NAMES = ["James","Mary","John","Patricia","Robert","Jennifer","Michael","Linda",
               "William","Barbara","David","Susan","Richard","Jessica","Thomas","Sarah",
               "Ji","Min","So","Yun","Hee","Sung","Ji-Young","Hyun","Jae","Soo","Bo"]
CITIES = ["Los Angeles","Torrance","Gardena","Koreatown","Rowland Heights","Cerritos",
          "Fullerton","Irvine","Pasadena","Glendale","Burbank","Alhambra"]
STATES = ["CA","NY","TX","WA","NJ","VA","IL","GA","FL","MA"]
LANGUAGES = ["Korean","English","Spanish","Vietnamese","Tagalog","Chinese","Japanese"]
PAYERS = ["Blue Shield","Kaiser","Anthem","Molina","LA Care","Health Net",
          "United Healthcare","Aetna","Cigna","Humana"]
PLAN_TYPES = ["HMO","PPO","EPO","POS","DHMO","Medicare Advantage","Medi-Cal"]
CPT_CODES = ["99213","99214","99215","99203","99204","93000","85025","80053",
             "71046","93306","43239","45378","99232","99233"]
ICD_CODES = ["I10","E11.9","Z00.00","J06.9","M54.5","K21.0","E78.5","F32.9",
             "Z87.39","I25.10","N18.3","E03.9","M17.11","J45.20"]
MEDICATIONS = ["Metformin 500mg","Lisinopril 10mg","Atorvastatin 20mg","Amlodipine 5mg",
               "Omeprazole 20mg","Metoprolol 25mg","Losartan 50mg","Gabapentin 300mg",
               "Levothyroxine 50mcg","Sertraline 50mg","Albuterol inhaler","Furosemide 20mg"]
PHARMACIES = ["CVS Pharmacy","Walgreens","Rite Aid","H Mart Pharmacy","Koreatown Pharmacy",
              "Target Pharmacy","Costco Pharmacy","Longs Drug"]
LABS = [("Hemoglobin A1c","LOINC:4548-4","5.5","4.0-5.6","%","normal"),
        ("Fasting Glucose","LOINC:1558-6","95","70-99","mg/dL","normal"),
        ("Total Cholesterol","LOINC:2093-3","210","<200","mg/dL","H"),
        ("Creatinine","LOINC:2160-0","1.1","0.6-1.2","mg/dL","normal"),
        ("TSH","LOINC:3016-3","2.1","0.4-4.0","mIU/L","normal"),
        ("Sodium","LOINC:2951-2","139","136-145","mEq/L","normal"),
        ("CBC WBC","LOINC:6690-2","7.2","4.0-11.0","K/uL","normal")]
PCP_PROVIDERS = [("Dr. James Park","1234567890","Internal Medicine","Seoul Medical Group","(310)555-0100"),
                 ("Dr. Susan Lee","2345678901","Family Medicine","LA Korean Clinic","(213)555-0200"),
                 ("Dr. David Kim","3456789012","Internal Medicine","SMG Torrance","(310)555-0300"),
                 ("Dr. Maria Gonzalez","4567890123","Family Medicine","Harbor Medical","(310)555-0400")]


def rand_date(start_year=1940, end_year=2005):
    start = date(start_year, 1, 1)
    end = date(end_year, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))


def rand_future_date(days_ahead=30, window=180):
    d = date.today() + timedelta(days=days_ahead + random.randint(0, window))
    return d.strftime("%Y-%m-%d")


def rand_past_date(days_back=365, window=730):
    d = date.today() - timedelta(days=random.randint(0, days_back + window))
    return d.strftime("%Y-%m-%d")


def make_patient_id():
    return f"PT{random.randint(100000, 999999)}"


# ── Row generators ─────────────────────────────────────────────────────────────

def gen_patients_rows(n, base_patients=None):
    """Generate patient rows. Returns (rows, patient_list)."""
    patients = []
    rows = []
    for _ in range(n):
        pid = make_patient_id()
        last = random.choice(ALL_SURNAMES)
        first = random.choice(FIRST_NAMES)
        dob = rand_date(1940, 1985).strftime("%Y-%m-%d")
        gender = random.choice(["M","F"])
        phone = f"({random.randint(200,999)}){random.randint(200,999)}-{random.randint(1000,9999)}"
        city = random.choice(CITIES)
        state = random.choice(STATES)
        lang = random.choice(LANGUAGES)
        patients.append({"patient_id": pid, "last_name": last, "first_name": first, "dob": dob})
        rows.append({
            "patient_id": pid,
            "last_name": last,
            "first_name": first,
            "dob": dob,
            "gender": gender,
            "phone": phone,
            "email": f"{first.lower()}.{last.lower()}@email.com",
            "address": f"{random.randint(100,9999)} {random.choice(['Main','Oak','Maple','Park','Wilshire'])} St",
            "city": city,
            "state": state,
            "zip": f"{random.randint(90001,96162)}",
            "language": lang,
            "assigned_broker": f"BRK{random.randint(1000,9999)}",
        })
    return rows, patients


def gen_eligibility_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        eff = rand_past_date(365, 365)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "payer_name": random.choice(PAYERS),
            "plan_name": f"{random.choice(PLAN_TYPES)} Gold",
            "member_id": f"MBR{random.randint(100000,999999)}",
            "group_number": f"GRP{random.randint(1000,9999)}",
            "effective_date": eff,
            "term_date": rand_future_date(90, 365),
            "status": random.choice(["Active","Active","Active","Inactive","Pending"]),
            "plan_type": random.choice(PLAN_TYPES),
            "copay": random.choice([10,20,30,40,50]),
            "deductible": random.choice([500,1000,1500,2000,3000]),
            "verified_date": rand_past_date(30, 60),
        })
    return rows


def gen_claims_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        billed = round(random.uniform(100, 5000), 2)
        allowed = round(billed * random.uniform(0.4, 0.9), 2)
        paid = round(allowed * random.uniform(0.7, 1.0), 2)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "claim_number": f"CLM{random.randint(1000000,9999999)}",
            "dos": rand_past_date(30, 365),
            "cpt_code": random.choice(CPT_CODES),
            "icd_codes": random.choice(ICD_CODES),
            "provider_name": random.choice([p[0] for p in PCP_PROVIDERS]),
            "provider_npi": f"{random.randint(1000000000,9999999999)}",
            "billed_amount": billed,
            "allowed_amount": allowed,
            "paid_amount": paid,
            "patient_responsibility": round(billed - paid, 2),
            "status": random.choice(["Paid","Paid","Paid","Denied","Pending","Adjusted"]),
            "denial_reason": random.choice(["","","","Not medically necessary","Auth required",""]),
            "submission_date": rand_past_date(60, 300),
            "paid_date": rand_past_date(30, 60),
        })
    return rows


def gen_pharmacy_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        fill = rand_past_date(7, 90)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "medication_name": random.choice(MEDICATIONS),
            "ndc_code": f"{random.randint(10000,99999)}-{random.randint(100,999)}-{random.randint(10,99)}",
            "dosage": random.choice(["500mg","10mg","20mg","25mg","50mg","100mg"]),
            "quantity": random.choice([30,60,90]),
            "days_supply": random.choice([30,60,90]),
            "pharmacy_name": random.choice(PHARMACIES),
            "pharmacy_phone": f"({random.randint(200,999)}){random.randint(200,999)}-{random.randint(1000,9999)}",
            "fill_date": fill,
            "refill_due_date": rand_future_date(20, 90),
            "refills_remaining": random.randint(0, 11),
            "status": random.choice(["Filled","Filled","Filled","On Hold","Expired"]),
            "last_fill_status": random.choice(["Dispensed","Dispensed","Returned","Partial"]),
        })
    return rows


def gen_labs_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        test = random.choice(LABS)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "test_name": test[0],
            "test_code": test[1],
            "result_value": test[2],
            "reference_range": test[3],
            "unit": test[4],
            "flag": test[5],
            "ordered_by": random.choice([p[0] for p in PCP_PROVIDERS]),
            "ordering_npi": f"{random.randint(1000000000,9999999999)}",
            "collection_date": rand_past_date(7, 180),
            "result_date": rand_past_date(5, 175),
            "lab_name": random.choice(["Quest Diagnostics","LabCorp","UCLA Lab","Cedars Lab"]),
            "status": random.choice(["Final","Final","Preliminary","Corrected"]),
        })
    return rows


def gen_authorizations_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "auth_number": f"AUTH{random.randint(1000000,9999999)}",
            "auth_type": random.choice(["Specialist Referral","Procedure","Inpatient","Outpatient"]),
            "service_type": random.choice(["Cardiology","Orthopedics","MRI","Physical Therapy","Surgery"]),
            "referring_provider": random.choice([p[0] for p in PCP_PROVIDERS]),
            "rendering_provider": random.choice([p[0] for p in PCP_PROVIDERS]),
            "requested_date": rand_past_date(30, 180),
            "approved_date": rand_past_date(25, 175),
            "start_date": rand_past_date(20, 60),
            "end_date": rand_future_date(30, 180),
            "approved_units": random.randint(1, 20),
            "used_units": random.randint(0, 10),
            "status": random.choice(["Approved","Approved","Denied","Pending","Expired"]),
            "denial_reason": random.choice(["","","Not medically necessary","Auth not requested",""]),
        })
    return rows


def gen_medications_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "medication_name": random.choice(MEDICATIONS),
            "ndc_code": f"{random.randint(10000,99999)}-{random.randint(100,999)}-{random.randint(10,99)}",
            "dosage": random.choice(["500mg","10mg","20mg","25mg","50mg"]),
            "frequency": random.choice(["Once daily","Twice daily","Three times daily","As needed"]),
            "quantity": random.choice([30,60,90]),
            "days_supply": random.choice([30,60,90]),
            "prescriber_name": random.choice([p[0] for p in PCP_PROVIDERS]),
            "prescriber_npi": f"{random.randint(1000000000,9999999999)}",
            "prescribed_date": rand_past_date(30, 365),
            "status": random.choice(["Active","Active","Discontinued","On Hold"]),
        })
    return rows


def gen_pcp_rows(n, patients):
    rows = []
    for _ in range(n):
        p = random.choice(patients)
        pcp = random.choice(PCP_PROVIDERS)
        rows.append({
            "patient_id": p["patient_id"],
            "last_name": p["last_name"],
            "dob": p["dob"],
            "pcp_name": pcp[0],
            "provider_npi": pcp[1],
            "specialty": pcp[2],
            "practice_name": pcp[3],
            "practice_phone": pcp[4],
            "practice_address": f"{random.randint(100,9999)} Wilshire Blvd, Los Angeles, CA",
            "assigned_date": rand_past_date(30, 730),
            "status": random.choice(["Active","Active","Inactive"]),
        })
    return rows


SHEET_GENERATORS = {
    "patients":       gen_patients_rows,
    "eligibility":    gen_eligibility_rows,
    "claims":         gen_claims_rows,
    "pharmacy":       gen_pharmacy_rows,
    "labs":           gen_labs_rows,
    "authorizations": gen_authorizations_rows,
    "medications":    gen_medications_rows,
    "pcp":            gen_pcp_rows,
}

ALL_TYPES = list(SHEET_GENERATORS.keys())


# ── File builders ──────────────────────────────────────────────────────────────

def build_excel_openpyxl(sheets: dict, filepath: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    wb.save(filepath)


def build_excel_xlsxwriter(sheets: dict, filepath: str):
    workbook = xlsxwriter.Workbook(filepath)
    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        ws = workbook.add_worksheet(sheet_name[:31])
        headers = list(rows[0].keys())
        for col, h in enumerate(headers):
            ws.write(0, col, h)
        for row_idx, row in enumerate(rows, 1):
            for col, h in enumerate(headers):
                ws.write(row_idx, col, row.get(h, ""))
    workbook.close()


def build_csv(sheets: dict, filepath: str):
    """For CSV, merge all sheets into one file (patients sheet takes priority)."""
    chosen = None
    for preferred in ["patients","claims","pharmacy","eligibility","labs","authorizations","medications","pcp"]:
        if preferred in sheets and sheets[preferred]:
            chosen = preferred
            break
    if not chosen:
        return

    rows = sheets[chosen]
    headers = list(rows[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def build_file(sheets: dict, filepath: str):
    if EXCEL_ENGINE == "openpyxl":
        build_excel_openpyxl(sheets, filepath)
    elif EXCEL_ENGINE == "xlsxwriter":
        build_excel_xlsxwriter(sheets, filepath)
    else:
        build_csv(sheets, filepath)


def file_extension():
    return ".csv" if EXCEL_ENGINE == "csv" else ".xlsx"


# ── Upload ─────────────────────────────────────────────────────────────────────

def get_auth_token(server_url: str, username: str, password: str) -> str | None:
    """Log in and return a Bearer token, or None on failure."""
    if not HAS_REQUESTS:
        return None
    try:
        r = requests.post(
            f"{server_url}/api/auth/login",
            json={"username": username, "password": password},
            timeout=15,
        )
        if r.status_code == 200:
            return r.json().get("token")
        print(f"[AUTH] Login failed ({r.status_code}): {r.json().get('error','?')}")
        return None
    except Exception as e:
        print(f"[AUTH] Login error: {e}")
        return None


def upload_batch(filepaths: list, server_url: str, token: str = None, timeout: int = 120) -> dict:
    if not HAS_REQUESTS:
        return {"error": "requests not installed", "uploaded": 0}

    files_payload = []
    file_handles = []
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        for fp in filepaths:
            fh = open(fp, "rb")
            file_handles.append(fh)
            mime = "text/csv" if fp.endswith(".csv") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            files_payload.append(("files", (os.path.basename(fp), fh, mime)))

        t0 = time.time()
        resp = requests.post(f"{server_url}/api/upload", files=files_payload, headers=headers, timeout=timeout)
        elapsed = time.time() - t0
        resp.raise_for_status()
        data = resp.json()
        data["_elapsed_s"] = round(elapsed, 2)
        return data
    except requests.exceptions.ConnectionError:
        return {"error": f"Cannot connect to {server_url} — is the server running?"}
    except requests.exceptions.Timeout:
        return {"error": f"Upload timed out after {timeout}s"}
    except Exception as e:
        return {"error": str(e)}
    finally:
        for fh in file_handles:
            fh.close()


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="SMG Bridge upload stress tester")
    parser.add_argument("--files",  type=int, default=20,   help="Number of Excel files to generate (default: 20)")
    parser.add_argument("--rows",   type=int, default=50,   help="Rows per sheet per file (default: 50)")
    parser.add_argument("--batch",  type=int, default=10,   help="Files per upload batch (default: 10, max: 200)")
    parser.add_argument("--url",    type=str, default="http://localhost:3000", help="Server base URL")
    parser.add_argument("--types",  type=str, default="all", help="Comma-separated sheet types to include (default: all)")
    parser.add_argument("--keep",   action="store_true",    help="Keep generated files after upload")
    parser.add_argument("--no-upload", action="store_true", help="Generate files only, do not upload")
    parser.add_argument("--out-dir", type=str, default="stress_test_files", help="Output directory for generated files")
    parser.add_argument("--user",   type=str, default=None, help="Username for auth (e.g. admin)")
    parser.add_argument("--pass",   type=str, dest="password", default=None, help="Password for auth (e.g. smgadmin)")
    args = parser.parse_args()

    # Validate batch size
    args.batch = min(args.batch, 200)

    # Resolve sheet types
    if args.types == "all":
        active_types = ALL_TYPES
    else:
        active_types = [t.strip() for t in args.types.split(",") if t.strip() in SHEET_GENERATORS]
        if not active_types:
            print(f"[ERROR] No valid types. Choose from: {', '.join(ALL_TYPES)}")
            sys.exit(1)

    # Ensure patients is always generated (needed as FK for other types)
    if "patients" not in active_types:
        active_types = ["patients"] + active_types

    out_dir = Path(args.out_dir)
    out_dir.mkdir(exist_ok=True)

    ext = file_extension()
    print(f"\n{'='*60}")
    print(f"  SMG Bridge — Upload Stress Tester")
    print(f"{'='*60}")
    print(f"  Files to generate : {args.files}")
    print(f"  Rows per sheet    : {args.rows}")
    print(f"  Batch size        : {args.batch}")
    print(f"  Sheet types       : {', '.join(active_types)}")
    print(f"  File format       : {ext.lstrip('.')} (engine: {EXCEL_ENGINE})")
    print(f"  Server            : {args.url}")
    print(f"  Output dir        : {out_dir.resolve()}")
    print(f"{'='*60}\n")

    # ── Step 1: Generate files ─────────────────────────────────────────────────
    print(f"[1/2] Generating {args.files} files...")
    generated = []
    t_gen_start = time.time()

    for i in range(1, args.files + 1):
        filepath = str(out_dir / f"stress_{i:04d}{ext}")

        # Always generate patients first (needed as source for other sheets)
        patient_rows, patients = gen_patients_rows(args.rows)
        sheets = {}

        for stype in active_types:
            if stype == "patients":
                sheets["patients"] = patient_rows
            else:
                gen_fn = SHEET_GENERATORS[stype]
                sheets[stype] = gen_fn(args.rows, patients)

        build_file(sheets, filepath)
        generated.append(filepath)

        size_kb = os.path.getsize(filepath) / 1024
        print(f"  [{i:04d}/{args.files}] {os.path.basename(filepath)}  ({size_kb:.1f} KB)", end="\r")

    t_gen = time.time() - t_gen_start
    print(f"\n  Generated {len(generated)} files in {t_gen:.1f}s\n")

    if args.no_upload or not HAS_REQUESTS:
        print(f"[SKIP] Upload skipped. Files are in: {out_dir.resolve()}")
        return

    # ── Authenticate if credentials provided ──────────────────────────────────
    token = None
    if args.user and args.password:
        print(f"[AUTH] Logging in as {args.user}...")
        token = get_auth_token(args.url, args.user, args.password)
        if token:
            print(f"[AUTH] Token acquired.\n")
        else:
            print(f"[AUTH] Login failed — uploads will proceed without auth (requires BRIDGE_AUTH_ENABLED=false on server).\n")

    # ── Step 2: Upload in batches ──────────────────────────────────────────────
    print(f"[2/2] Uploading in batches of {args.batch}...")
    batches = [generated[i:i+args.batch] for i in range(0, len(generated), args.batch)]

    total_uploaded = 0
    total_patients = 0
    total_errors = 0
    batch_times = []
    upload_errors = []

    t_upload_start = time.time()

    for b_idx, batch in enumerate(batches, 1):
        print(f"  Batch {b_idx}/{len(batches)} ({len(batch)} files)...", end=" ", flush=True)
        result = upload_batch(batch, args.url, token=token)

        if "error" in result:
            print(f"FAILED — {result['error']}")
            upload_errors.append(f"Batch {b_idx}: {result['error']}")
            total_errors += len(batch)
            if "Cannot connect" in result.get("error",""):
                print("\n[FATAL] Server not reachable. Start it with:  npm start")
                break
        else:
            elapsed = result.get("_elapsed_s", 0)
            uploaded = result.get("uploaded", 0)
            patients_now = result.get("patientCount", 0)
            batch_times.append(elapsed)
            total_uploaded += uploaded
            total_patients = patients_now

            file_results = result.get("results", [])
            file_errors = [r for r in file_results if not r.get("success")]
            total_errors += len(file_errors)

            print(f"OK  {elapsed:.1f}s  |  patients in DB: {patients_now:,}")

    t_upload = time.time() - t_upload_start

    # ── Summary ────────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  STRESS TEST SUMMARY")
    print(f"{'='*60}")
    print(f"  Files generated   : {len(generated)}")
    print(f"  Files uploaded    : {total_uploaded}")
    print(f"  Upload errors     : {total_errors}")
    print(f"  Patients in DB    : {total_patients:,}")
    print(f"  Gen time          : {t_gen:.1f}s")
    print(f"  Upload time       : {t_upload:.1f}s")
    if batch_times:
        print(f"  Avg batch time    : {sum(batch_times)/len(batch_times):.2f}s")
        print(f"  Fastest batch     : {min(batch_times):.2f}s")
        print(f"  Slowest batch     : {max(batch_times):.2f}s")
    if upload_errors:
        print(f"\n  Errors:")
        for e in upload_errors:
            print(f"    • {e}")
    print(f"{'='*60}\n")

    # ── Cleanup ────────────────────────────────────────────────────────────────
    if not args.keep:
        removed = 0
        for fp in generated:
            try:
                os.remove(fp)
                removed += 1
            except Exception:
                pass
        if out_dir.exists() and not list(out_dir.iterdir()):
            out_dir.rmdir()
        print(f"  Cleaned up {removed} temporary files.")
    else:
        print(f"  Files kept in: {out_dir.resolve()}")

    print()


if __name__ == "__main__":
    main()
